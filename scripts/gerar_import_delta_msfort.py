# -*- coding: utf-8 -*-
"""
Delta MSFORT: compara MSFORT_GESTÃO (1).xlsx (antigo) com MSFORT_ATUAL.xlsx (novo)
e gera SQL só com os REGISTROS NOVOS (fornecedores, despesas->pagar, pagamentos->receber),
já classificados (categoria+subcategoria). Direcionado à empresa MSFORT.
Uso: python scripts/gerar_import_delta_msfort.py -> supabase/import_msfort_delta.sql
"""
import openpyxl, uuid, io, unicodedata, datetime
OLD = "MSFORT_GESTÃO (1).xlsx"; NEW = "MSFORT_ATUAL.xlsx"; OUT = "supabase/import_msfort_delta.sql"
NS = uuid.UUID("11111111-2222-3333-4444-555555555555")
def uid(e, l): return str(uuid.uuid5(NS, f"{e}:{l}"))
def norm(s): return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().upper().strip()
def q(v):
    if v is None: return "null"
    t = str(v).strip()
    return "null" if t == "" or t.lower() == "none" else "'" + t.replace("'", "''") + "'"
def digits(v):
    if v is None: return "null"
    t = str(v).strip()
    if t.endswith(".0"): t = t[:-2]
    d = "".join(c for c in t if c.isdigit())
    return "'" + d + "'" if d else "null"
def num(v):
    if v is None or str(v).strip() == "": return "null"
    try: return str(round(float(v), 2))
    except: return "null"
def dd(v):
    if isinstance(v, (datetime.datetime, datetime.date)): return "'" + v.strftime("%Y-%m-%d") + "'"
    t = str(v or "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try: return "'" + datetime.datetime.strptime(t, fmt).strftime("%Y-%m-%d") + "'"
        except: pass
    return "null"

MAP_DESP = {
    "VARIAVEL": "Custo de material / insumos", "FORNECEDORES": "Custo de material / insumos",
    "COMPRA DE MATERIAL": "Custo de material / insumos", "MATERIAL DA OBRA": "Custo de material / insumos",
    "COMPRA DE E.P.I": "Custo de material / insumos", "MAO DE OBRA": "Custo de mão de obra direta",
    "FOLHA DE PAGAMENTO": "Despesas com pessoal (folha/encargos)", "PRO-LABORE": "Despesas com pessoal (folha/encargos)",
    "PROLABORE": "Despesas com pessoal (folha/encargos)", "ADIANTAMENTO": "Despesas com pessoal (folha/encargos)",
    "RECISAO DE CONTRATO": "Despesas com pessoal (folha/encargos)", "RECISAO AVULSO": "Despesas com pessoal (folha/encargos)",
    "DESPESAS MEDICAS": "Despesas com pessoal (folha/encargos)", "COMISSOES": "Despesas comerciais / marketing",
    "ADMINISTRATIVAS": "Despesas administrativas", "DESPESA MARLEY": "Despesas administrativas",
    "ALIMENTACAO": "Despesas administrativas", "TELEFONIA": "Despesas administrativas", "FIXA": "Despesas administrativas",
    "TAXAS": "Impostos e taxas", "CARTAO DE CREDITO": "Juros e tarifas bancárias",
    "TRANSPORTE": "Manutenção e frota", "FRETES": "Manutenção e frota", "LOCACAO DE AUTOMOVEL": "Manutenção e frota",
    "ATIVO CIRCULANTE": "Investimentos / imobilizado", "EMPRESTIMO": "Outras despesas",
    "FINANCIAMENTO": "Outras despesas", "OUTROS": "Outras despesas",
}
DESP_DEFAULT = "Outras despesas"

def rows(fn, sheet):
    wb = openpyxl.load_workbook(fn, data_only=True, read_only=True)
    n = [s for s in wb.sheetnames if norm(s) == norm(sheet)][0]
    data = list(wb[n].iter_rows(values_only=True)); hdr = None; out = []
    for r in data:
        if any(c is not None and str(c).strip() for c in r):
            if hdr is None: hdr = [str(c).strip() if c is not None else "" for c in r]
            else: out.append(dict(zip(hdr, r)))
    return out

def novos(sheet, idcol):
    o = {str(r.get(idcol)).strip() for r in rows(OLD, sheet) if r.get(idcol)}
    return [r for r in rows(NEW, sheet) if r.get(idcol) and str(r.get(idcol)).strip() not in o]

# referências do arquivo NOVO
clientes_ok = {str(r.get("ID_Cliente")).strip() for r in rows(NEW, "Clientes") if r.get("ID_Cliente")}
forn_por_nome = {}
for r in rows(NEW, "Fornecedores"):
    nm = norm(r.get("Nome"))
    if nm and r.get("ID_Fornecedor"): forn_por_nome[nm] = uid("fornecedor", str(r.get("ID_Fornecedor")).strip())

out = io.StringIO(); w = out.write
w("-- IMPORT MSFORT — DELTA (só registros novos). Direcionado à MSFORT. Idempotente.\n\n")
w("do $$\ndeclare v_tenant uuid; v_cat uuid;\nbegin\n")
w("  select id into v_tenant from empresas_consultoras where nome = 'MSFORT' limit 1;\n")
w("  if v_tenant is null then raise exception 'Empresa MSFORT nao encontrada'; end if;\n")
w("  perform seed_categorias_dre(v_tenant);  -- garante categorias\n")
w("  perform set_config('app.pular_baixa', 'on', true);\n\n")

# fornecedores novos
nf = novos("Fornecedores", "ID_Fornecedor")
w(f"  -- fornecedores novos ({len(nf)})\n")
for r in nf:
    w(f"  insert into fornecedores (id, empresa_consultora_id, nome, documento) values "
      f"('{uid('fornecedor', str(r.get('ID_Fornecedor')).strip())}', v_tenant, {q(r.get('Nome'))}, {digits(r.get('CNPJ'))}) on conflict (id) do nothing;\n")

# coleta subcategorias necessárias dos novos títulos
subcats = {}  # su_uuid -> (cat_nome, sub_nome)
def registra_sub(cat, subraw):
    sub = str(subraw).strip().title() if subraw and str(subraw).strip() else "Geral"
    su = uid("subcat", f"{cat}|{sub.upper()}")
    subcats[su] = (cat, sub); return su

nd = novos("Despesas", "ID_Despesa"); npg = novos("Pagamentos", "ID_Pagamento")
def title_despesa(r):
    valor = num(r.get("Valor")); idd = r.get("ID_Despesa")
    tid = uid("despesa", str(idd).strip())
    cat = MAP_DESP.get(norm(r.get("Categoria_Despesa")), DESP_DEFAULT)
    su = registra_sub(cat, r.get("Categoria_Despesa"))
    forn_nome = str(r.get("Forncedor") or "").strip()
    forn = f"'{forn_por_nome[norm(forn_nome)]}'" if norm(forn_nome) in forn_por_nome else "null"
    desc = str(r.get("Descrição") or "").strip() or (forn_nome or "Despesa")
    nf_ = r.get("NumeroNF")
    if nf_ and str(nf_).strip(): desc += f" (NF {str(nf_).strip().rstrip('.0')})"
    pago = norm(r.get("Status")) == "PAGO"
    return (tid, "pagar", desc, "null", forn, cat, su, valor,
            dd(r.get("Data_Vencimento")), dd(r.get("Data_Despesa")),
            "pago" if pago else "aberto", dd(r.get("Data_Despesa")) if pago else "null")
def title_pag(r):
    valor = num(r.get("Valor")); idp = r.get("ID_Pagamento")
    tid = uid("pagamento", str(idp).strip())
    emp = norm(r.get("TipoReceita")) == "EMPRESTIMO"
    cat = "Outras receitas" if emp else "Receita de serviços"
    su = registra_sub(cat, "Empréstimo" if emp else (r.get("TipoReceita") or "Serviços"))
    idc = str(r.get("ID_Cliente")).strip() if r.get("ID_Cliente") else None
    cli = f"'{uid('cliente', idc)}'" if idc in clientes_ok else "null"
    desc = str(r.get("Servico_Produto") or r.get("Observacoes") or "Recebimento").strip()
    nf_ = r.get("NF")
    if nf_ and str(nf_).strip(): desc += f" (NF {str(nf_).strip().rstrip('.0')})"
    pago = norm(r.get("Status_Pagamento")) == "PAGO"
    return (tid, "receber", desc, cli, "null", cat, su, valor,
            dd(r.get("Data_Vencimento")), dd(r.get("Data")),
            "pago" if pago else "aberto", dd(r.get("Data")) if pago else "null")

titulos = [title_despesa(r) for r in nd if num(r.get("Valor")) != "null"]
titulos += [title_pag(r) for r in npg if num(r.get("Valor")) != "null"]

# subcategorias (garante que existam)
w(f"\n  -- subcategorias necessárias ({len(subcats)})\n")
for su, (cat, sub) in subcats.items():
    w(f"  insert into subcategorias_financeiras (id, empresa_consultora_id, categoria_id, nome) "
      f"select '{su}', v_tenant, c.id, {q(sub)} from categorias_financeiras c "
      f"where c.empresa_consultora_id = v_tenant and c.nome = {q(cat)} on conflict (id) do nothing;\n")

# títulos novos
w(f"\n  -- títulos novos ({len(titulos)})\n")
for (tid, tipo, desc, cli, forn, cat, su, valor, venc, comp, status, pgto) in titulos:
    w(f"  insert into titulos_financeiros (id, empresa_consultora_id, tipo, descricao, cliente_id, fornecedor_id, "
      f"categoria_id, subcategoria_id, valor, vencimento, competencia, status, data_pagamento, origem) values "
      f"('{tid}', v_tenant, '{tipo}', {q(desc)}, {cli}, {forn}, "
      f"(select id from categorias_financeiras where empresa_consultora_id=v_tenant and nome={q(cat)}), '{su}', "
      f"{valor}, {venc}, {comp}, '{status}', {pgto}, 'manual') on conflict (id) do nothing;\n")

w("\nend $$;\n")
w(f"\n-- fornecedores novos={len(nf)} despesas novas={len(nd)} pagamentos novos={len(npg)} titulos={len(titulos)}\n")
with open(OUT, "w", encoding="utf-8") as f: f.write(out.getvalue())
print(f"OK -> {OUT}")
print(f"fornecedores={len(nf)} despesas={len(nd)} pagamentos={len(npg)} titulos={len(titulos)} subcats={len(subcats)}")
