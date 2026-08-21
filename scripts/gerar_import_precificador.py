# -*- coding: utf-8 -*-
"""
Import precificador: Parametros -> parametros_preco ; ProdutosServicos -> campos
de preço em produtos ; ComposicaoCustos -> composicao_custo (por categoria).
Requer 0021 + fase 1 (produtos). Idempotente.
Uso: python scripts/gerar_import_precificador.py -> supabase/import_msfort_precificador.sql
"""
import openpyxl, uuid, io, unicodedata
XLSX = "MSFORT_GESTÃO (1).xlsx"; OUT = "supabase/import_msfort_precificador.sql"
NS = uuid.UUID("11111111-2222-3333-4444-555555555555")
def uid(e, l): return str(uuid.uuid5(NS, f"{e}:{l}"))
def norm(s): return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().upper().strip()
def q(v): return "'" + str(v).replace("'", "''") + "'"
def n(v):
    try: return round(float(v), 4)
    except: return 0.0

wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
def rows(nn):
    x = [s for s in wb.sheetnames if norm(s) == nn][0]; data = list(wb[x].iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in data[0]]
    return hdr, [dict(zip(hdr, r)) for r in data[1:] if any(c is not None and str(c).strip() for c in r)]

# categoria a partir da descrição
def categoria(desc):
    d = norm(desc)
    if any(k in d for k in ["MAO DE OBRA", "M.O", "HOMEM HORA", "FABRICA", "MONTAGEM", "SOLDA", "PINTURA", "CORTE"]): return "mao_de_obra"
    if any(k in d for k in ["FRETE", "TRANSPORTE", "LOGISTICA"]): return "frete"
    if any(k in d for k in ["MAQUINA", "LIXADEIRA", "PLATAFORMA", "GERADOR", "LOCACAO", "ANDAIME", "EQUIP"]): return "equipamento"
    if any(k in d for k in ["TERCEIR", "SUBCONTRAT"]): return "terceiro"
    if any(k in d for k in ["EPI", "SEGURANCA", "ALIMENTA", "TAXA"]): return "outros"
    return "material"

# Parametros
hp, PA = rows("PARAMETROS")
par = {norm(r.get(hp[0])): n(r.get(hp[1])) for r in PA}
overhead_hora = par.get("OVERHEAD_HORA_R$", 0) or par.get("OVERHEAD_HORA_R", 0)
overhead_perc = par.get("OVERHEADER_%", 0) or par.get("OVERHEAD_%", 0)
taxa_cartao = par.get("TAXA_CARTAO_%", 0)
meta = par.get("FATURAMENTO_META_MES", 0)

# ProdutosServicos + valid ids
hps, PS = rows("PRODUTOSSERVICOS")
prod_ids = {str(r.get("ID_Prod")).strip() for r in PS if r.get("ID_Prod")}

out = io.StringIO(); w = out.write
w("-- IMPORT MSFORT — precificador. Requer 0021 + fase 1. Idempotente.\n\n")
w("do $$\ndeclare v_tenant uuid;\nbegin\n  select id into v_tenant from empresas_consultoras limit 1;\n\n")

w("  -- parâmetros\n")
w(f"  insert into parametros_preco (id, empresa_consultora_id, overhead_metodo, overhead_hora, overhead_perc, taxa_cartao_perc, meta_faturamento) "
  f"values ('{uid('parpreco','unico')}', v_tenant, 'hora', {overhead_hora}, {overhead_perc}, {taxa_cartao}, {meta}) "
  f"on conflict (empresa_consultora_id) do update set overhead_hora=excluded.overhead_hora, overhead_perc=excluded.overhead_perc, "
  f"taxa_cartao_perc=excluded.taxa_cartao_perc, meta_faturamento=excluded.meta_faturamento;\n\n")

w("  -- campos de preço nos produtos\n")
for r in PS:
    idp = str(r.get("ID_Prod")).strip() if r.get("ID_Prod") else ""
    if not idp: continue
    marg = n(r.get("Margem_Alvo_%")); tempo = n(r.get("TempoHora")); lista = n(r.get("Preco_Lista"))
    cat = r.get("Categoria")
    w(f"  update produtos set margem_alvo={marg}, tempo_horas={tempo}, preco_lista={lista}, categoria={q(str(cat).strip()) if cat else 'null'} "
      f"where id='{uid('produto', idp)}' and empresa_consultora_id=v_tenant;\n")

w("\n  -- composição de custo (itens diretos)\n")
hc, CC = rows("COMPOSICAOCUSTOS")
ncomp = 0
for r in CC:
    idp = str(r.get("ID_Prod")).strip() if r.get("ID_Prod") else ""
    if idp not in prod_ids: continue
    if norm(r.get("Tipo")) == "PERCENTUAL": continue   # taxas tratadas nos parâmetros
    desc = r.get("Descricao")
    if not desc or not str(desc).strip(): continue
    cid = uid("comp", str(r.get("ID_Comp")).strip()) if r.get("ID_Comp") else uid("comp", f"{idp}|{desc}")
    w(f"  insert into composicao_custo (id, empresa_consultora_id, produto_id, categoria, descricao, quantidade, custo_unitario) "
      f"values ('{cid}', v_tenant, '{uid('produto', idp)}', '{categoria(desc)}', {q(str(desc).strip())}, "
      f"{n(r.get('Qtd_Base')) or 1}, {n(r.get('Custo_Unit'))}) on conflict (id) do nothing;\n")
    ncomp += 1

w("\nend $$;\n")
w(f"\n-- params(overhead_hora={overhead_hora}, taxa_cartao={taxa_cartao}, meta={meta}) ; {len(PS)} produtos ; {ncomp} itens de composição\n")
with open(OUT, "w", encoding="utf-8") as f: f.write(out.getvalue())
print(f"OK -> {OUT}")
print(f"overhead_hora={overhead_hora} taxa_cartao={taxa_cartao} meta={meta} produtos={len(PS)} composicao={ncomp}")
