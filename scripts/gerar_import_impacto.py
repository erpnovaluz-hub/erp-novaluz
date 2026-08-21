# -*- coding: utf-8 -*-
"""
Import Impacto: CadPecas -> pecas ; Produção -> producao.
Resolve cliente_id/colaborador_id (uuid5 da fase 1) e peca_id (por nome).
Requer: 0014 aplicada + fase 1 importada. Idempotente.
Uso: python scripts/gerar_import_impacto.py -> supabase/import_msfort_impacto.sql
"""
import openpyxl, uuid, io, unicodedata, datetime

XLSX = "MSFORT_GESTÃO (1).xlsx"
OUT  = "supabase/import_msfort_impacto.sql"
NS   = uuid.UUID("11111111-2222-3333-4444-555555555555")

def uid(e, l): return str(uuid.uuid5(NS, f"{e}:{l}"))
def norm(s): return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().upper()
def s(v):
    if v is None: return "null"
    t = str(v).strip()
    return "null" if t == "" or t.lower() == "none" else "'" + t.replace("'", "''") + "'"
def num(v):
    if v is None or str(v).strip() == "": return "0"
    try: return str(round(float(v), 4))
    except: return "0"
def d(v):
    if isinstance(v, (datetime.datetime, datetime.date)): return "'" + v.strftime("%Y-%m-%d") + "'"
    t = str(v or "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try: return "'" + datetime.datetime.strptime(t, fmt).strftime("%Y-%m-%d") + "'"
        except: pass
    return "null"

wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
def sheet(nome_norm):
    n = [x for x in wb.sheetnames if norm(x) == nome_norm][0]
    return wb[n]
def rows_of(ws):
    data = list(ws.iter_rows(values_only=True)); hdr = None; start = 0
    for i, r in enumerate(data):
        if any(c is not None and str(c).strip() for c in r):
            hdr = [str(c).strip() if c is not None else "" for c in r]; start = i + 1; break
    for r in data[start:]:
        if any(c is not None and str(c).strip() for c in r): yield hdr, dict(zip(hdr, r))
def col(hdr, *subs):
    for h in hdr:
        if all(sub in norm(h) for sub in subs): return h
    return None

# referências da fase 1
clientes_ok = {str(r.get("ID_Cliente")).strip() for _, r in rows_of(sheet("CLIENTES")) if r.get("ID_Cliente")}
func_ws = sheet("FUNCIONARIOS")
func_ok = set()
for hdr, r in rows_of(func_ws):
    c = col(hdr, "ID", "FUNCION")
    if c and r.get(c): func_ok.add(str(r.get(c)).strip())

out = io.StringIO(); w = out.write
w("-- ============================================================================\n")
w("-- IMPORT MSFORT — IMPACTO: pecas + producao. Idempotente. Rode no SQL Editor.\n")
w("-- Requer: 0014 + fase 1.\n")
w("-- ============================================================================\n\n")
w("do $$\ndeclare v_tenant uuid;\nbegin\n")
w("  select id into v_tenant from empresas_consultoras limit 1;\n\n")

# ---- pecas ----
pecas_nomes = set()
npc = 0
w("  -- peças\n")
for hdr, r in rows_of(sheet("CADPECAS")):
    cn = col(hdr, "NOME"); cp = col(hdr, "PESO"); cv = col(hdr, "VALOR")
    nome = r.get(cn) if cn else None
    if not nome or not str(nome).strip(): continue
    nome = str(nome).strip()
    pecas_nomes.add(nome)
    w(f"  insert into pecas (id, empresa_consultora_id, nome, peso, valor_kg) values "
      f"('{uid('peca', nome)}', v_tenant, {s(nome)}, {num(r.get(cp))}, {num(r.get(cv))}) on conflict (id) do nothing;\n")
    npc += 1

# ---- produção (em lotes) ----
w("\n  -- produção\n")
ws = sheet("PRODUCAO")
prod_rows = list(rows_of(ws))
hdr0 = prod_rows[0][0] if prod_rows else []
c_id   = col(hdr0, "ID", "PRODUC")
c_cli  = col(hdr0, "ID", "CLIENTE")
c_fun  = col(hdr0, "ID", "FUNCION")
c_cat  = col(hdr0, "CATEGORIA")
c_pec  = col(hdr0, "PECA") or col(hdr0, "PE", "A")
c_qtd  = col(hdr0, "QUANTIDADE")
c_pu   = col(hdr0, "PESO", "UNIT")
c_vu   = col(hdr0, "VALOR", "UNIT")
c_tipo = col(hdr0, "TIPO")
c_data = col(hdr0, "DATA")

lote, batch, nprod = [], 200, 0
def flush():
    global lote
    if not lote: return
    w("  insert into producao (id, empresa_consultora_id, data, cliente_id, colaborador_id, peca_id, peca_nome, categoria, tipo, quantidade, peso_unit, valor_unit) values\n")
    w("    " + ",\n    ".join(lote) + "\n  on conflict (id) do nothing;\n")
    lote = []

for hdr, r in prod_rows:
    idp = r.get(c_id)
    if not idp: continue
    cli = str(r.get(c_cli)).strip() if r.get(c_cli) else ""
    fun = str(r.get(c_fun)).strip() if r.get(c_fun) else ""
    pnome = str(r.get(c_pec)).strip() if r.get(c_pec) else ""
    cli_sql = f"'{uid('cliente', cli)}'" if cli in clientes_ok else "null"
    fun_sql = f"'{uid('colab', fun)}'" if fun in func_ok else "null"
    pec_sql = f"'{uid('peca', pnome)}'" if pnome in pecas_nomes else "null"
    lote.append(
        f"('{uid('prod', str(idp).strip())}', v_tenant, {d(r.get(c_data))}, {cli_sql}, {fun_sql}, {pec_sql}, "
        f"{s(pnome)}, {s(r.get(c_cat))}, {s(r.get(c_tipo))}, {num(r.get(c_qtd))}, {num(r.get(c_pu))}, {num(r.get(c_vu))})"
    )
    nprod += 1
    if len(lote) >= batch: flush()
flush()

w("\nend $$;\n")
w(f"\n-- {npc} peças, {nprod} lançamentos de produção\n")
with open(OUT, "w", encoding="utf-8") as f: f.write(out.getvalue())
print(f"OK -> {OUT}")
print(f"pecas={npc} producao={nprod}")
