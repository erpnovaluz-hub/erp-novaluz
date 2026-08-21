# -*- coding: utf-8 -*-
"""
Import bônus: Tabela_bonus -> bonus_regras (por tipo) ; Bonus -> bonus_producao.
Requer 0017 + fase 1. Idempotente.
Uso: python scripts/gerar_import_bonus.py -> supabase/import_msfort_bonus.sql
"""
import openpyxl, uuid, io, unicodedata, datetime
from collections import Counter

XLSX = "MSFORT_GESTÃO (1).xlsx"; OUT = "supabase/import_msfort_bonus.sql"
NS = uuid.UUID("11111111-2222-3333-4444-555555555555")
def uid(e, l): return str(uuid.uuid5(NS, f"{e}:{l}"))
def norm(s): return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().upper().strip()
def n2(v):
    try: return round(float(v), 2)
    except: return 0.0
def d(v):
    if isinstance(v, (datetime.datetime, datetime.date)): return v.strftime("%Y-%m-%d")
    t = str(v or "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try: return datetime.datetime.strptime(t, fmt).strftime("%Y-%m-%d")
        except: pass
    return None

wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
def rows(nn):
    n = [x for x in wb.sheetnames if norm(x) == nn][0]; data = list(wb[n].iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in data[0]]
    return hdr, [dict(zip(hdr, r)) for r in data[1:] if any(c is not None and str(c).strip() for c in r)]
def col(hdr, *subs):
    for h in hdr:
        if all(sub in norm(h) for sub in subs): return h
    return None

func_ok = set()
for r in rows("FUNCIONARIOS")[1]:
    for k, v in r.items():
        if "ID" in norm(k) and "FUNC" in norm(k) and v: func_ok.add(str(v).strip())

# regras por tipo (moda dos valores)
ht, TB = rows("TABELA_BONUS")
c_tipo = ht[1]; c_min = col(ht, "MINIMO"); c_fixo = col(ht, "FIXO"); c_p50 = col(ht, "50")
por_tipo = {}
agg = {}
for r in TB:
    tp = str(r.get(c_tipo) or "").strip().upper()
    if tp not in ("LD", "LP", "LPP"): continue
    agg.setdefault(tp, Counter())[(n2(r.get(c_min)), n2(r.get(c_fixo)), n2(r.get(c_p50)))] += 1
for tp, cnt in agg.items():
    por_tipo[tp] = cnt.most_common(1)[0][0]

out = io.StringIO(); w = out.write
w("-- IMPORT MSFORT — bônus (regras + produção diária). Idempotente. Requer 0017 + fase 1.\n\n")
w("do $$\ndeclare v_tenant uuid;\nbegin\n  select id into v_tenant from empresas_consultoras limit 1;\n\n")
w("  -- regras por tipo\n")
for tp, (mn, fx, p50) in por_tipo.items():
    w(f"  insert into bonus_regras (id, empresa_consultora_id, tipo, minimo, bonus_fixo, bonus_por_50) values "
      f"('{uid('bonusregra', tp)}', v_tenant, '{tp}', {mn}, {fx}, {p50}) on conflict (id) do nothing;\n")

# produção diária
hb, B = rows("BONUS")
c_f = col(hb, "FUNCION"); c_dt = col(hb, "DATA")
c_ld = col(hb, "LD"); c_lp = [c for c in hb if norm(c) == "LP"]; c_lpp = col(hb, "LPP"); c_na = col(hb, "NA")
c_lp = c_lp[0] if c_lp else None
w("\n  -- produção diária de bônus\n")
lote, nprod = [], 0
def flush():
    global lote
    if not lote: return
    w("  insert into bonus_producao (id, empresa_consultora_id, data, colaborador_id, ld, lp, lpp, na) values\n    "
      + ",\n    ".join(lote) + "\n  on conflict (id) do nothing;\n")
    lote = []
for r in B:
    f = str(r.get(c_f) or "").strip(); dt = d(r.get(c_dt))
    if not f or not dt or f not in func_ok: continue
    bid = uid("bonusprod", f"{dt}|{f}")
    lote.append(f"('{bid}', v_tenant, '{dt}', '{uid('colab', f)}', {n2(r.get(c_ld))}, {n2(r.get(c_lp))}, {n2(r.get(c_lpp))}, {n2(r.get(c_na))})")
    nprod += 1
    if len(lote) >= 200: flush()
flush()

w("\nend $$;\n")
w(f"\n-- regras: {por_tipo} ; {nprod} dias de produção de bônus\n")
with open(OUT, "w", encoding="utf-8") as f: f.write(out.getvalue())
print(f"OK -> {OUT}")
print("regras:", por_tipo, "producao:", nprod)
