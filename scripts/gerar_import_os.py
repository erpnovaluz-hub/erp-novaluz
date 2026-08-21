# -*- coding: utf-8 -*-
"""
Import OS: PlanosAcao -> ordens_servico ; Atividades -> atividades_os.
Resolve cliente_id/colaborador_id pelos uuids da fase 1. Requer 0016 + fase 1.
Uso: python scripts/gerar_import_os.py -> supabase/import_msfort_os.sql
"""
import openpyxl, uuid, io, unicodedata, datetime

XLSX = "MSFORT_GESTÃO (1).xlsx"; OUT = "supabase/import_msfort_os.sql"
NS = uuid.UUID("11111111-2222-3333-4444-555555555555")
def uid(e, l): return str(uuid.uuid5(NS, f"{e}:{l}"))
def norm(s): return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().upper().strip()
def s(v):
    if v is None: return "null"
    t = str(v).strip()
    return "null" if t == "" or t.lower() == "none" else "'" + t.replace("'", "''") + "'"
def n2(v):
    try: return str(round(float(v), 2))
    except: return "0"
def pct(v):
    try: return str(round(float(v) * 100, 2))
    except: return "0"
def d(v):
    if isinstance(v, (datetime.datetime, datetime.date)): return "'" + v.strftime("%Y-%m-%d") + "'"
    t = str(v or "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try: return "'" + datetime.datetime.strptime(t, fmt).strftime("%Y-%m-%d") + "'"
        except: pass
    return "null"

wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
def rows(nn):
    n = [x for x in wb.sheetnames if norm(x) == nn][0]; data = list(wb[n].iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in data[0]]
    return [dict(zip(hdr, r)) for r in data[1:] if any(c is not None and str(c).strip() for c in r)]

clientes_ok = {str(r.get("ID_Cliente")).strip() for r in rows("CLIENTES") if r.get("ID_Cliente")}
func_ok = set()
for r in rows("FUNCIONARIOS"):
    for k, v in r.items():
        if "ID" in norm(k) and "FUNC" in norm(k) and v: func_ok.add(str(v).strip())

URG = {"ALTO": "alta", "ALTA": "alta", "MEDIO": "media", "MEDIA": "media", "BAIXO": "baixa", "BAIXA": "baixa"}
ST_OS = {"A FAZER": "a_fazer", "EM ANDAMENTO": "em_andamento", "CONCLUIDO": "concluido", "CANCELADO": "cancelado"}
ST_AT = {"CONCLUIDO": "concluido", "NAO INICIADO": "nao_iniciado", "EM ANDAMENTO": "em_andamento", "PARADO": "parado"}

out = io.StringIO(); w = out.write
w("-- IMPORT MSFORT — OS (ordens_servico + atividades_os). Idempotente.\n")
w("-- Requer 0016 + fase 1.\n\ndo $$\ndeclare v_tenant uuid;\nbegin\n")
w("  select id into v_tenant from empresas_consultoras limit 1;\n\n")

nos = 0; os_ids = set()
w("  -- ordens de serviço\n")
for r in rows("PLANOSACAO"):
    ida = r.get("ID_Acao")
    if not ida: continue
    ida = str(ida).strip()
    os_ids.add(ida)
    titulo = r.get("Oque") if r.get("Oque") and str(r.get("Oque")).strip() else f"OS {ida}"
    cli = str(r.get("ID_Cliente")).strip() if r.get("ID_Cliente") else ""
    cli_sql = f"'{uid('cliente', cli)}'" if cli in clientes_ok else "null"
    urg = URG.get(norm(r.get("Urgencia")))
    st = ST_OS.get(norm(r.get("Status_Acao")), "a_fazer")
    w("  insert into ordens_servico (id, empresa_consultora_id, cliente_id, titulo, motivo, local, como_sera_feito, "
      "responsavel, prazo, data_realizado, custo_estimado, urgencia, status) values "
      f"('{uid('os', ida)}', v_tenant, {cli_sql}, {s(titulo)}, {s(r.get('Porque'))}, "
      f"{s(r.get('Onde'))}, {s(r.get('ComoSeraFeito'))}, {s(r.get('Responsável') or r.get('Responsavel'))}, "
      f"{d(r.get('Prazo'))}, {d(r.get('Data_realizado'))}, {n2(r.get('CustoEstimado'))}, "
      f"{s(urg) if urg else 'null'}, '{st}') on conflict (id) do nothing;\n")
    nos += 1

nat = 0
w("\n  -- atividades\n")
for r in rows("ATIVIDADES"):
    idt = r.get("ID_Atividade"); ida = r.get("ID_Acao")
    if not idt or not ida or not r.get("Descricao"): continue
    if str(ida).strip() not in os_ids: continue   # atividade órfã (OS inexistente)
    resp = str(r.get("Responsavel")).strip() if r.get("Responsavel") else ""
    colab = f"'{uid('colab', resp)}'" if resp in func_ok else "null"
    st = ST_AT.get(norm(r.get("Status")), "nao_iniciado")
    w("  insert into atividades_os (id, empresa_consultora_id, os_id, descricao, colaborador_id, setor, "
      "data_inicio, data_fim, status, conclusao_pct, alocacao_pct) values "
      f"('{uid('ativ', str(idt).strip())}', v_tenant, '{uid('os', str(ida).strip())}', {s(r.get('Descricao'))}, "
      f"{colab}, {s(r.get('Setor'))}, {d(r.get('Data_Inicio'))}, {d(r.get('Data_Fim'))}, '{st}', "
      f"{pct(r.get('Conclusao_%'))}, {pct(r.get('Alocacao_%'))}) on conflict (id) do nothing;\n")
    nat += 1

w("\nend $$;\n")
w(f"\n-- {nos} ordens de serviço, {nat} atividades\n")
with open(OUT, "w", encoding="utf-8") as f: f.write(out.getvalue())
print(f"OK -> {OUT}  os={nos} atividades={nat}")
