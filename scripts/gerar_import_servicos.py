# -*- coding: utf-8 -*-
"""
Import serviços: CadServico -> servicos ; preenche producao.servico_id + unidade.
Ao setar unidade, o valor_total (gerado) recalcula ciente de KG/UND.
Requer: 0014 + 0015 + import impacto. Idempotente.
Uso: python scripts/gerar_import_servicos.py -> supabase/import_msfort_servicos.sql
"""
import openpyxl, uuid, io, unicodedata
from collections import defaultdict

XLSX = "MSFORT_GESTÃO (1).xlsx"
OUT  = "supabase/import_msfort_servicos.sql"
NS   = uuid.UUID("11111111-2222-3333-4444-555555555555")

def uid(e, l): return str(uuid.uuid5(NS, f"{e}:{l}"))
def norm(s): return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().upper()
def q(v): return "'" + str(v).replace("'", "''") + "'"
def num(v):
    try: return str(round(float(v), 4))
    except: return "0"

wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
def rows(nome_norm):
    n = [x for x in wb.sheetnames if norm(x) == nome_norm][0]; data = list(wb[n].iter_rows(values_only=True))
    hdr = [str(c).strip() if c is not None else "" for c in data[0]]
    return hdr, [dict(zip(hdr, r)) for r in data[1:] if any(c is not None and str(c).strip() for c in r)]
def col(hdr, *subs):
    for h in hdr:
        if all(sub in norm(h) for sub in subs): return h
    return None

hs, CS = rows("CADSERVICO")
c_id = col(hs, "ID", "SERVICO"); c_nome = col(hs, "NOME"); c_uni = col(hs, "UNIDADE"); c_val = col(hs, "VALOR")
servico_unidade = {}  # id_legado -> 'KG'/'UND'
serv_validos = set()

out = io.StringIO(); w = out.write
w("-- ============================================================================\n")
w("-- IMPORT MSFORT — serviços + ligação na produção. Idempotente.\n")
w("-- Requer: 0014 + 0015 + import impacto.\n")
w("-- ============================================================================\n\n")
w("do $$\ndeclare v_tenant uuid;\nbegin\n")
w("  select id into v_tenant from empresas_consultoras limit 1;\n\n")
w("  -- serviços\n")
for r in CS:
    ids = str(r.get(c_id)).strip() if r.get(c_id) else ""
    nome = r.get(c_nome)
    if not ids or not nome or not str(nome).strip(): continue
    uni = (str(r.get(c_uni) or "KG").strip().upper() or "KG")
    uni = "UND" if uni.startswith("UN") else "KG"
    servico_unidade[ids] = uni
    serv_validos.add(ids)
    w(f"  insert into servicos (id, empresa_consultora_id, nome, unidade, valor) values "
      f"('{uid('servico', ids)}', v_tenant, {q(str(nome).strip())}, '{uni}', {num(r.get(c_val))}) on conflict (id) do nothing;\n")

# backfill producao: agrupa prod-ids por serviço
hp, P = rows("PRODUCAO")
c_pid = col(hp, "ID", "PRODUC"); c_serv = col(hp, "SERVI")
por_serv = defaultdict(list)
for r in P:
    pid = r.get(c_pid); sid = str(r.get(c_serv)).strip() if r.get(c_serv) else ""
    if pid and sid in serv_validos:
        por_serv[sid].append(uid("prod", str(pid).strip()))

w("\n  -- liga serviço + unidade na produção (recalcula valor_total)\n")
for sid, ids in por_serv.items():
    su = uid("servico", sid); uni = servico_unidade[sid]
    for i in range(0, len(ids), 300):
        lista = ",".join(f"'{x}'" for x in ids[i:i+300])
        w(f"  update producao set servico_id = '{su}', unidade = '{uni}' "
          f"where empresa_consultora_id = v_tenant and id in ({lista});\n")

w("\nend $$;\n")
w(f"\n-- {len(serv_validos)} serviços, {sum(len(v) for v in por_serv.values())} produções ligadas\n")
with open(OUT, "w", encoding="utf-8") as f: f.write(out.getvalue())
print(f"OK -> {OUT}")
print(f"servicos={len(serv_validos)} producoes_ligadas={sum(len(v) for v in por_serv.values())}")
