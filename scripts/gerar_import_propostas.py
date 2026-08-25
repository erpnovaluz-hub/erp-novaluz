# -*- coding: utf-8 -*-
"""
Importa propostas + itens da MSFORT (MSFORT_ATUAL.xlsx) e da T4 (ERP_T4 INDUSTRIAL.xlsx)
para o novo sistema, cada uma no seu tenant. Idempotente. UUIDs: MSFORT sem prefixo,
T4 com prefixo 't4:' (mesmo esquema dos imports anteriores).
Uso: python scripts/gerar_import_propostas.py -> supabase/import_propostas.sql
"""
import openpyxl, uuid, io, unicodedata, datetime
NS = uuid.UUID("11111111-2222-3333-4444-555555555555")
OUT = "supabase/import_propostas.sql"
def norm(s): return unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().upper().strip()
def q(v):
    if v is None: return "null"
    t = str(v).strip()
    return "null" if t == "" or t.lower() == "none" else "'" + t.replace("'", "''") + "'"
def num(v):
    if v is None or str(v).strip() == "": return "null"
    try: return str(round(float(v), 2))
    except: return "null"
def parsedate(v):
    if isinstance(v, (datetime.datetime, datetime.date)): return v if isinstance(v, datetime.date) else v.date()
    t = str(v or "").strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try: return datetime.datetime.strptime(t, fmt).date()
        except: pass
    return None
def dsql(dt): return "'" + dt.strftime("%Y-%m-%d") + "'" if dt else "null"

def rows(fn, sheet):
    wb = openpyxl.load_workbook(fn, data_only=True, read_only=True)
    n = [s for s in wb.sheetnames if norm(s) == norm(sheet)][0]
    data = list(wb[n].iter_rows(values_only=True)); hdr = None; out = []
    for r in data:
        if any(c is not None and str(c).strip() for c in r):
            if hdr is None: hdr = [str(c).strip() if c is not None else "" for c in r]
            else: out.append(dict(zip(hdr, r)))
    return out

out = io.StringIO(); w = out.write
w("-- IMPORT PROPOSTAS + ITENS — MSFORT e T4. Idempotente. Cada uma no seu tenant.\n\n")

# ================= MSFORT =================
def uid_ms(e, l): return str(uuid.uuid5(NS, f"{e}:{l}"))
ST_MS = {"ENVIADA": "enviada", "RASCUNHO": "rascunho", "ACEITA": "aprovada", "ADITIVO": "enviada", "RECUSADA": "recusada"}
FN = "MSFORT_ATUAL.xlsx"
cli_ok = {str(r.get("ID_Cliente")).strip() for r in rows(FN, "Clientes") if r.get("ID_Cliente")}
props = rows(FN, "Propostas"); itens = rows(FN, "ItensProposta")
prop_ids = set()
w("do $$\ndeclare v_tenant uuid;\nbegin\n")
w("  select id into v_tenant from empresas_consultoras where nome = 'MSFORT' limit 1;\n")
w("  if v_tenant is null then raise exception 'MSFORT nao encontrada'; end if;\n\n  -- propostas MSFORT\n")
npm = 0
for r in props:
    idp = str(r.get("ID_Proposta") or "").strip()
    if not idp: continue
    prop_ids.add(idp)
    cidr = str(r.get("ID_Cliente") or "").strip()
    cli = f"'{uid_ms('cliente', cidr)}'" if cidr in cli_ok else "null"
    st = ST_MS.get(norm(r.get("Status_Proposta")), "rascunho")
    demi = parsedate(r.get("Data_Proposta"))
    val = None
    try: val = demi + datetime.timedelta(days=int(float(r.get("Validade")))) if demi and r.get("Validade") else None
    except: val = None
    w("  insert into propostas (id, empresa_consultora_id, cliente_id, numero, data, validade, status, objeto, escopo_desc, premissas, pagamento) values "
      f"('{uid_ms('proposta', idp)}', v_tenant, {cli}, {q(r.get('Numero'))}, {dsql(demi)}, {dsql(val)}, '{st}', "
      f"{q(r.get('Objeto_da_Proposta'))}, {q(r.get('Especificaoes_Tecnicas_do_Escopo'))}, {q(r.get('Condicoes_e_Exclusoes_da_Proposta'))}, {q(r.get('Condicoes_Pagamento'))}) on conflict (id) do nothing;\n")
    npm += 1
w("\n  -- itens MSFORT\n")
nim = 0
for r in itens:
    idi = str(r.get("ID_Item") or "").strip(); idp = str(r.get("ID_Proposta") or "").strip()
    if not idi or idp not in prop_ids: continue
    desc = r.get("Descricao") or r.get("Produto_Servico") or "Item"
    w("  insert into itens_proposta (id, empresa_consultora_id, proposta_id, descricao, referencia, quantidade, valor_unit) values "
      f"('{uid_ms('itemprop', idi)}', v_tenant, '{uid_ms('proposta', idp)}', {q(desc)}, {q(r.get('ID_Prod'))}, "
      f"coalesce({num(r.get('Quant'))},1), coalesce({num(r.get('Preco_Unit'))},0)) on conflict (id) do nothing;\n")
    nim += 1
w("end $$;\n\n")

# ================= T4 =================
def uid_t4(e, l): return str(uuid.uuid5(NS, f"t4:{e}:{l}"))
FT = "ERP_T4 INDUSTRIAL.xlsx"
# clientes válidos (Pessoas com Tipo_Relacao Cliente)
t4_cli_ok = {str(r.get("ID_Pessoa")).strip() for r in rows(FT, "Pessoas_Clientes_Fornecedores") if r.get("ID_Pessoa") and norm(r.get("Tipo_Relacao")) == "CLIENTE"}
cat_nome = {str(r.get("ID_Item")).strip(): r.get("Nome_Item") for r in rows(FT, "Catalogo_Produtos_Servicos") if r.get("ID_Item")}
tprops = rows(FT, "Propostas"); titens = rows(FT, "Propostas_Itens")
tprop_ids = set()
w("do $$\ndeclare v_tenant uuid;\nbegin\n")
w("  select id into v_tenant from empresas_consultoras where nome = 'T4 INDUSTRIAL' limit 1;\n")
w("  if v_tenant is null then raise exception 'T4 INDUSTRIAL nao encontrada'; end if;\n\n  -- propostas T4\n")
npt = 0
for i, r in enumerate(sorted(tprops, key=lambda x: str(parsedate(x.get("Data_Emissao")) or "")), 1):
    idp = str(r.get("ID_Proposta") or "").strip()
    if not idp: continue
    tprop_ids.add(idp)
    cidr = str(r.get("ID_Cliente") or "").strip()
    cli = f"'{uid_t4('cliente', cidr)}'" if cidr in t4_cli_ok else "null"
    demi = parsedate(r.get("Data_Emissao")); dval = parsedate(r.get("Data_Validade"))
    ano = demi.year if demi else 2026
    numero = f"PROP-{ano}-{i:04d}"
    st = "enviada"  # 'Em Aberto'
    w("  insert into propostas (id, empresa_consultora_id, cliente_id, numero, data, validade, status, apresentacao, premissas, pagamento) values "
      f"('{uid_t4('proposta', idp)}', v_tenant, {cli}, '{numero}', {dsql(demi)}, {dsql(dval)}, '{st}', "
      f"{q(r.get('Escopo_Intro'))}, {q(r.get('Premissas'))}, {q(r.get('Cond_Pagamento'))}) on conflict (id) do nothing;\n")
    npt += 1
w("\n  -- itens T4\n")
nit = 0
for r in titens:
    idi = str(r.get("ID_Proposta_Item") or "").strip(); idp = str(r.get("ID_Proposta") or "").strip()
    if not idi or idp not in tprop_ids: continue
    iditem = str(r.get("ID_Item") or "").strip()
    desc = cat_nome.get(iditem) or "Item"
    w("  insert into itens_proposta (id, empresa_consultora_id, proposta_id, descricao, referencia, quantidade, valor_unit) values "
      f"('{uid_t4('itemprop', idi)}', v_tenant, '{uid_t4('proposta', idp)}', {q(desc)}, {q(iditem)}, "
      f"coalesce({num(r.get('Quantidade'))},1), coalesce({num(r.get('Preco_Unitario'))},0)) on conflict (id) do nothing;\n")
    nit += 1
w("end $$;\n")
w(f"\n-- MSFORT: {npm} propostas, {nim} itens ; T4: {npt} propostas, {nit} itens\n")
with open(OUT, "w", encoding="utf-8") as f: f.write(out.getvalue())
print(f"OK -> {OUT}")
print(f"MSFORT propostas={npm} itens={nim} ; T4 propostas={npt} itens={nit}")
